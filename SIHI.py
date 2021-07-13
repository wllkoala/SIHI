import math
import os
import traceback
from sys import exit
from tkinter import Tk, filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

root = Tk()
root.withdraw()


class ATO_check():
    def __init__(self, ato_report, storage_report):
        self.ato_report = ato_report
        self.storage_report = storage_report

    def start_to_run(self):
        self.import_ato_report()
        self.import_storage_report()
        self.df_buffer = self.df_storage_report_ATO.copy()
        self.get_bom()
        self.save_ato_excel()

    def import_ato_report(self):
        self.df_ato = pd.read_excel(self.ato_report,
                                    sheet_name="ATO",
                                    engine="openpyxl")
        self.df_ato = self.df_ato.loc[:, ~self.df_ato.columns.str.
                                      contains('Unnamed')]
        self.df_bom_list = pd.read_excel(ato_report,
                                         sheet_name="BomList",
                                         engine="openpyxl")
        self.df_bom_list = self.df_bom_list.loc[:, ~self.df_bom_list.columns.
                                                str.contains('Unnamed')]
        print(self.df_bom_list)
        print(self.df_ato)

    def import_storage_report(self):
        df_storage_report = pd.read_excel(self.storage_report,
                                          engine="openpyxl")
        df_storage_report.fillna(0, inplace=True)
        non_use_sub = [
            "DALIAN-FG", "DALIAN-RAW", "LC-FPD", "LOL-FPD", "MRO-FPD",
            "QA-INSP", "QA-MRB", "QD-FG", "QD-RAW", "RT-W", "SY-FG", "SY-RAW",
            "TOOL-FPD", "Tooling", "RT-V"
        ]
        df_storage_report_ATO = df_storage_report.copy()
        for sub in non_use_sub:
            df_storage_report_ATO = df_storage_report_ATO[
                df_storage_report_ATO["Sub"] != sub]
        df_storage_report_ATO = df_storage_report_ATO[
            df_storage_report_ATO["Project"] == 0]
        df_storage_report_ATO = df_storage_report_ATO.drop(
            ["Org", "Sub", "Locator", "Project"], axis=1)
        self.df_storage_report_ATO = df_storage_report_ATO.groupby(
            "Item").sum()
        print(self.df_storage_report_ATO)

    def get_bom(self):
        self.df_ato["BOM"] = "-"
        for index, sn in enumerate(self.df_ato["SN"]):
            self.df_bom = self.df_bom_list.loc[self.df_bom_list["SN"] == sn, :]
            if not self.df_bom.empty:
                print(sn)
                self.cal_avaliable_qty(index, sn)
                self.df_ato.loc[index, "Avaliable"] = self.avaliable
                sheet_name = f'{index}_{sn}'
                self.df_ato.loc[
                    index, "BOM"] = f'=HYPERLINK("#{sheet_name}!A1","BOM")'
            else:
                self.df_ato.loc[index, "BOM"] = "NO BOM!"
        self.df_ato.fillna(0, inplace=True)
        self.df_ato["Diff"] = self.df_ato["Qty"] - self.df_ato["Avaliable"]
        print(self.df_ato)

    def cal_avaliable_qty(self, index, sn):
        self.df_bom = self.df_bom.drop(["SN"], axis=1)
        self.df_bom_description = self.df_bom.loc[:, ["Item", "Description"]]
        self.df_bom = self.df_bom.groupby("Item").sum()
        self.df_bom_avaliable = self.df_bom.copy()
        self.df_bom[
            "Qty_need"] = self.df_ato["Qty"].iloc[index] * self.df_bom["Qty"]
        df_cal_qty = pd.merge(self.df_bom,
                              self.df_buffer,
                              on="Item",
                              how='left')
        df_cal_qty.fillna(0, inplace=True)
        df_cal_qty["diff"] = (df_cal_qty["On-hand"] -
                              df_cal_qty["Qty_need"]) / self.df_bom["Qty"]
        df_cal_qty = pd.merge(self.df_bom_description,
                              df_cal_qty,
                              on="Item",
                              how='right')
        df_cal_qty[
            "Back To Summary"] = '=HYPERLINK("#Summary!A1","Back To Summary")'
        print(df_cal_qty)
        if df_cal_qty["diff"].min() >= 0:
            self.avaliable = self.df_ato["Qty"].iloc[index]
        else:
            self.avaliable = math.floor(self.df_ato["Qty"].iloc[index] +
                                        df_cal_qty["diff"].min())
            if self.avaliable < 0:
                self.avaliable = 0
        print(df_cal_qty["diff"].min())
        print("---")
        print(self.avaliable)
        self.df_bom_avaliable[
            "Qty"] = self.avaliable * self.df_bom_avaliable["Qty"]
        self.df_buffer = pd.merge(self.df_buffer,
                                  self.df_bom_avaliable,
                                  on="Item",
                                  how='left')
        self.df_buffer.fillna(0, inplace=True)
        self.df_buffer[
            "On-hand"] = self.df_buffer["On-hand"] - self.df_buffer["Qty"]
        self.df_buffer = self.df_buffer.drop("Qty", axis=1)
        wb = load_workbook("Summary.xlsx")
        with pd.ExcelWriter("Summary.xlsx", engine="openpyxl") as writer:
            writer.book = wb
            df_cal_qty.to_excel(writer, sheet_name=f'{index}_{sn}', index=None)

    def save_ato_excel(self):
        # 装载excel
        wb = load_workbook("Summary.xlsx")
        if "Summary" in wb.sheetnames:
            wb.remove(wb["Summary"])
        # 如果有多个模块可以读写excel文件，这里要指定engine，否则可能会报错
        with pd.ExcelWriter("Summary.xlsx", engine="openpyxl") as writer:
            # 没有下面这个语句的话excel表将完全被覆盖
            writer.book = wb
            # 将df_summary写入Excel
            self.df_ato.to_excel(writer, sheet_name="Summary", index=True)
            self.move_sheet(wb)

    def move_sheet(self, wb, from_loc=None, to_loc=None):
        sheets = wb._sheets
        # if no from_loc given, assume last sheet
        if from_loc is None:
            from_loc = len(sheets) - 1
        # if no to_loc given, assume first
        if to_loc is None:
            to_loc = 0
        sheet = sheets.pop(from_loc)
        sheets.insert(to_loc, sheet)


if __name__ == "__main__":
    try:
        ato_report = filedialog.askopenfilename(title="ato_report",
                                                filetypes=[("All files", "*")])
        storage_report = filedialog.askopenfilename(title="storage_report",
                                                    filetypes=[("All files",
                                                                "*")])
        summary_workbook = Workbook()
        ws = summary_workbook.active
        ws.title = "Summary"
        summary_workbook.save("Summary.xlsx")
        ato = ATO_check(ato_report, storage_report)
        ato.start_to_run()
        input('Press <Enter>')
    except Exception as err:
        messagebox.showerror("Warning!", err)
        with open(os.path.join(os.getcwd(), "error.txt"), "w+") as f:
            traceback.print_exc(file=f)
        print(err)
        exit()
